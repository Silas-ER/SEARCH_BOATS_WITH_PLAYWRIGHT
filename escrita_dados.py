import openpyxl

def exportar(dados):
    # Abra o arquivo Excel existente
    workbook = openpyxl.load_workbook('dados_consultados.xlsx')

    # Escolha a planilha na qual você deseja trabalhar (substitua 'Planilha1' pelo nome da sua planilha)
    sheet = workbook['Locate']

    # Função para encontrar a primeira linha vazia na planilha
    def encontrar_primeira_linha_vazia(sheet):
        for row in range(1, sheet.max_row + 1):
            is_empty = all(sheet.cell(row=row, column=col).value is None for col in range(1, sheet.max_column + 1))
            if is_empty:
                return row
        return sheet.max_row + 1  # Se não houver linha vazia, insira na próxima linha vazia

    # Encontre a primeira linha vazia na planilha
    linha_vazia = encontrar_primeira_linha_vazia(sheet)
    coluna = 1

    for dado in dados:
        sheet.cell(row=linha_vazia, column=coluna, value=dado['data'])
        sheet.cell(row=linha_vazia, column=coluna + 1, value=dado['name'])
        sheet.cell(row=linha_vazia, column=coluna + 2, value=dado['latitude'])
        sheet.cell(row=linha_vazia, column=coluna + 3, value=dado['longitude'])
        linha_vazia += 1  # Vá para a próxima linha vazia

    # Salve o arquivo
    workbook.save('dados_consultados.xlsx')